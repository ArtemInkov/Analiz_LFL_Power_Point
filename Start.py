"""Импорт os для адаптации путей
Импорт pandas для работы с файлами эксель
xlsxwriter для записи xl файлов
openpyxl import load_workbook для чтение xl файлов
import xlrd Библиотека для чтения и форматирования данных xls или xlsx
import xlwt
from openpyxl.utils.dataframe import dataframe_to_rows"""
from class_shop import Shop
import os
from pptx import Presentation
from pptx.enum.chart import XL_CHART_TYPE
from pptx.chart.data import CategoryChartData
from pptx.util import Inches
from itertools import islice
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import xlsxwriter
from openpyxl import load_workbook
import xlrd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
import xlwt
from func_pptx import presentation_pptx_shops_graphic, presentation_pptx_ferst

print('Начало работы')

analize = xlwt.Workbook(encoding="utf-8") #Создаем переменную с книгой
sheet_analize_LFL = analize.add_sheet("Анализ год к году") #Создаем лист
data_sale = load_workbook('./2.xlsx') # Присваеваем файл с данными продаж к переменной data_sale


sheet = data_sale['TDSheet']
#sheet_data_sale = data_sale.wb.get_sheet_by_name('TDSheet') #Присваеваем лист  с данными продаж к переменной sheet_data_sale из файла data_sale
max_col = sheet.max_column  #записываем в переменную максимальное число колонок
max_Row = sheet.max_row #записываем в переменную максимальное число строк
last_coordinate = sheet.cell(row =sheet.max_row, column=sheet.max_column) # присваеваем координаты последней ячейки переменной
count = 0 # cчетчик столбцов для записи месяцов
shop_dict = {} # словарь ключ номер магазина, значение номер строки в файле с импортом данных
shop_dict_new = {} #словарь ключ номер магазина, значение номер строки в файле для экспорта данных данных
shop_class_dict = {}
"""Цикл проходит по всей таблице"""
for cellObj in sheet['A1':last_coordinate.coordinate]:
    for cell in cellObj:
        """Условный оператор if ищет начало таблицы по слову 'Магазин' и задает координаты
        далее записывает в новую таблицу заголовок магазины"""
        if cell.value == 'Магазин':
            #print(cell.column)
            #print(cell.row)
            sheet_analize_LFL.write(1, 0, cell.value)
            count = 3
            for i_index in range (cell.row + 1, max_Row):
                if sheet.cell(row=i_index, column= cell.column).value != None:
                    shop_split_list = sheet.cell(row=i_index, column= cell.column).value.split()
                    for number in shop_split_list:
                        if number.isdigit():
                            if int(number) in shop_dict:
                                shop_dict[int(number)].append(i_index)
                                shop_class_dict.get(int(number)).index_shop_list.append(i_index)
                            else:
                                shop_dict[int(number)] = []
                                shop_dict[int(number)].append(i_index)
                                shop = Shop(f'Краснодар {number} ')
                                shop_class_dict[int(number)] = shop
                                shop_class_dict.get(int(number)).index_shop_list.append(i_index)

            for key, value in sorted(shop_dict.items()):
                sheet_analize_LFL.write(count, 0, f'{key} Краснодар')
                shop_dict_new[count] = value
                count += 1
            """Цикл проходит по месяца """
            count = 0
            for i_index in range(2, max_col):
                """Условный оператор if фильтрует значения None и месяца 2022"""
                if sheet.cell(row=cell.row, column=i_index).value != None and not str(sheet.cell(row=cell.row, column=i_index).value).endswith('23'):
                    """Цикл продолжает идти по месяцам с места где остановился первый цикл в поисках совпадения 2023 года"""
                    for j_index in range(i_index + 1, max_col):
                        if str(sheet.cell(row=cell.row, column=j_index).value).startswith(sheet.cell(row=cell.row, column=i_index).value[0:4]):
                            count += 1
                            """Записывает месяц 2022 в  рублях """
                            #print('записываю', sheet.cell(row=cell.row, column=i_index).value + ' рубли')
                            sheet_analize_LFL.write(1, count, sheet.cell(row=cell.row, column=i_index).value + ' рубли')
                            count += 1
                            """Записывает следующей строкой месяц 2023 года в рублях  после месяца 2022 года"""
                            sheet_analize_LFL.write(1, count, sheet.cell(row=cell.row, column=j_index).value+ ' рубли')
                            count += 1
                            """Записывает месяц 2022 в вес """
                            sheet_analize_LFL.write(1, count, sheet.cell(row=cell.row, column=i_index).value + ' вес')
                            #print('Записываю', sheet.cell(row=cell.row, column=i_index).value + ' вес')
                            for i_row in range(cell.row +1, max_Row):
                                if sheet.cell(row=i_row, column=i_index).value != None:
                                    for key, value in shop_dict_new.items():
                                       #print(i_row, value)
                                       if i_row in value:
                                            #print('ok')
                                            #print(sheet.cell(row=i_row, column=i_index).value)
                                            #print('Записываю', sheet.cell(row=i_row, column=i_index).value)
                                            sheet_analize_LFL.write(key, count, sheet.cell(row=i_row, column=i_index).value)
                                            sheet_analize_LFL.write(key, count - 2,
                                                                    sheet.cell(row=i_row, column=i_index + 1).value)
                            count += 1
                            """Записывает следующей строкой месяц 2023 года в вес  после месяца 2022 года"""
                            #print('Записываю', sheet.cell(row=cell.row, column=j_index).value + ' вес')
                            sheet_analize_LFL.write(1, count, sheet.cell(row=cell.row, column=j_index).value + ' вес')
                            for i_row in range(cell.row +1, max_Row):
                                if sheet.cell(row=i_row, column=j_index).value != None:
                                    for key, value in shop_dict_new.items():
                                       #print(i_row, value, sheet.cell(row=i_row, column=j_index).value)
                                       if i_row in value:
                                            #print('ok')
                                            #print('записываю в',key, count )
                                            #print(sheet.cell(row=i_row, column=j_index).value)
                                            #print('Записываю', sheet.cell(row=i_row, column=j_index).value)
                                            sheet_analize_LFL.write(key, count, sheet.cell(row=i_row, column=j_index).value)
                                            sheet_analize_LFL.write(key, count - 2,
                                                                    sheet.cell(row=i_row, column=j_index + 1).value)
        analize.save("анализ продаж.xls")
        break




                #print(cell.row)
                #print(sheet.cell(row=7, column=4).value)
                #sheet_analize_LFL.write(0, cell.column, mounth)
#print(shop_dict)
#print(shop_dict_new)
#print(sorted(shop_dict.items()))
#
#print(shop_class_dict[2].index_shop_list)



a = 'анализ продаж.xlsx'
presentation_pptx_ferst(a)
presentation_pptx_shops_graphic(a)